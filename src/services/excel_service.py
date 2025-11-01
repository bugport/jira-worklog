"""Excel file operations for Jira work logs."""

from pathlib import Path
from typing import List, Optional, Tuple, Dict
from collections import defaultdict
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from rich.console import Console
from rich.progress import Progress, SpinnerColumn, TextColumn
from pydantic import ValidationError

from ..models.issue import Issue
from ..models.worklog import WorkLogEntry, SyncResult, ExistingWorkLog, WorkLogUpdate
from ..utils.formatters import format_date, parse_time_hours, parse_date
from ..utils.validators import validate_issue_key, validate_time_hours, validate_date
from decimal import Decimal

console = Console()


class ExcelService:
    """Service for Excel file operations."""
    
    REQUIRED_COLUMNS = [
        "Hierarchy Number",
        "Issue Key",
        "Summary",
        "Type",
        "Parent Issue Key",
        "Parent Issue Type",
        "Time Logged (hours)",
        "Date",
        "Comment",
        "Status"
    ]
    
    READ_ONLY_COLUMNS = ["Issue Key", "Summary", "Type"]
    
    def __init__(self):
        """Initialize Excel service."""
        pass
    
    def export_issues_to_excel(self, issues: List[Issue], output_file: str) -> bool:
        """Export issues to Excel template file.
        
        Args:
            issues: List of Issue objects
            output_file: Output Excel file path
            
        Returns:
            True if successful, False otherwise
        """
        try:
            if not issues:
                console.print("[yellow]No issues to export.[/yellow]")
                return False
            
            with Progress(
                SpinnerColumn(),
                TextColumn("[progress.description]{task.description}"),
                console=console
            ) as progress:
                task = progress.add_task("Creating Excel template...", total=None)
                
                # Create DataFrame from issues
                data = []
                for issue in issues:
                    data.append(issue.to_dict())
                
                df = pd.DataFrame(data)
                
                # Ensure all required columns exist
                for col in self.REQUIRED_COLUMNS:
                    if col not in df.columns:
                        df[col] = ""
                
                # Reorder columns
                df = df[self.REQUIRED_COLUMNS]
                
                progress.update(task, description="Writing Excel file...")
                
                # Write to Excel with formatting
                output_path = Path(output_file)
                output_path.parent.mkdir(parents=True, exist_ok=True)
                
                # Create Excel writer
                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Work Logs')
                    
                    # Get workbook and worksheet for formatting
                    workbook = writer.book
                    worksheet = writer.sheets['Work Logs']
                    
                    # Format header row
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF")
                    
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Format columns
                    worksheet.column_dimensions['A'].width = 15  # Issue Key
                    worksheet.column_dimensions['B'].width = 50  # Summary
                    worksheet.column_dimensions['C'].width = 15  # Type
                    worksheet.column_dimensions['D'].width = 18  # Time Logged
                    worksheet.column_dimensions['E'].width = 15  # Date
                    worksheet.column_dimensions['F'].width = 40  # Comment
                    worksheet.column_dimensions['G'].width = 15  # Status
                    
                    # Add data validation hints (using comments)
                    for row in range(2, len(df) + 2):
                        # Issue Key - read-only indicator
                        issue_key_cell = worksheet[f'A{row}']
                        if issue_key_cell.value:
                            issue_key_cell.font = Font(color="808080")  # Gray for read-only
                        
                        # Time Logged - decimal format hint
                        time_cell = worksheet[f'D{row}']
                        time_cell.number_format = '0.00'
                        
                        # Date - date format hint
                        date_cell = worksheet[f'E{row}']
                        date_cell.number_format = 'YYYY-MM-DD'
                    
                    # Add instructions sheet
                    instructions_df = pd.DataFrame({
                        'Column': [
                            'Issue Key',
                            'Summary',
                            'Type',
                            'Time Logged (hours)',
                            'Date',
                            'Comment',
                            'Status'
                        ],
                        'Description': [
                            'Jira issue key (read-only)',
                            'Issue summary (read-only)',
                            'Issue type (read-only)',
                            'Time logged in hours (decimal, e.g., 2.5)',
                            'Work log date (YYYY-MM-DD format)',
                            'Work log comment (optional)',
                            'Sync status (auto-populated)'
                        ],
                        'Required': [
                            'Yes',
                            'No',
                            'No',
                            'Yes',
                            'Yes',
                            'No',
                            'No'
                        ]
                    })
                    
                    instructions_df.to_excel(writer, index=False, sheet_name='Instructions')
                    
                    # Format instructions sheet
                    inst_worksheet = writer.sheets['Instructions']
                    for cell in inst_worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    inst_worksheet.column_dimensions['A'].width = 25
                    inst_worksheet.column_dimensions['B'].width = 60
                    inst_worksheet.column_dimensions['C'].width = 15
                
                progress.update(task, description=f"[green]Excel file created: {output_file}[/green]")
            
            console.print(f"[green]✓[/green] Exported {len(issues)} issue(s) to [cyan]{output_file}[/cyan]")
            console.print(f"[dim]Please fill in 'Time Logged (hours)', 'Date', and optionally 'Comment' columns.[/dim]")
            return True
            
        except Exception as e:
            console.print(f"[red]Error exporting to Excel:[/red] {str(e)}")
            return False
    
    def import_worklogs_from_excel(self, input_file: str) -> List[WorkLogEntry]:
        """Import work log entries from Excel file.
        
        Args:
            input_file: Input Excel file path
            
        Returns:
            List of WorkLogEntry objects
        """
        try:
            input_path = Path(input_file)
            if not input_path.exists():
                raise FileNotFoundError(f"Excel file not found: {input_file}")
            
            with Progress(
                SpinnerColumn(),
                TextColumn("[progress.description]{task.description}"),
                console=console
            ) as progress:
                task = progress.add_task("Reading Excel file...", total=None)
                
                # Read Excel file
                df = pd.read_excel(input_path, sheet_name='Work Logs')
                
                progress.update(task, description="Validating data...")
                
                # Validate required columns
                missing_columns = [col for col in ['Issue Key', 'Time Logged (hours)', 'Date'] if col not in df.columns]
                if missing_columns:
                    raise ValueError(f"Missing required columns: {', '.join(missing_columns)}")
                
                # Parse and validate rows
                entries = []
                errors = []
                
                for idx, row in df.iterrows():
                    try:
                        # Get values
                        issue_key = str(row['Issue Key']).strip() if pd.notna(row['Issue Key']) else ""
                        time_str = str(row['Time Logged (hours)']).strip() if pd.notna(row['Time Logged (hours)']) else ""
                        date_str = str(row['Date']).strip() if pd.notna(row['Date']) else ""
                        comment = str(row['Comment']).strip() if pd.notna(row.get('Comment')) and pd.notna(row['Comment']) else ""
                        
                        # Skip rows with empty required fields
                        if not issue_key or not time_str or not date_str:
                            continue
                        
                        # Validate issue key
                        if not validate_issue_key(issue_key):
                            errors.append(f"Row {idx + 2}: Invalid issue key format: {issue_key}")
                            continue
                        
                        # Parse time
                        try:
                            time_hours = parse_time_hours(time_str)
                        except ValueError as e:
                            errors.append(f"Row {idx + 2}: {str(e)}")
                            continue
                        
                        # Parse date
                        try:
                            # Handle different date formats
                            if isinstance(row['Date'], pd.Timestamp):
                                work_date = row['Date'].date()
                            elif isinstance(row['Date'], str):
                                work_date = parse_date(date_str)
                            else:
                                work_date = parse_date(date_str)
                        except ValueError as e:
                            errors.append(f"Row {idx + 2}: {str(e)}")
                            continue
                        
                        # Create work log entry (using alias 'date' for Excel column compatibility)
                        entry = WorkLogEntry(
                            issue_key=issue_key,
                            time_logged_hours=time_hours,
                            date=work_date,  # Using alias
                            comment=comment if comment else None
                        )
                        
                        entries.append(entry)
                        
                    except ValidationError as e:
                        errors.append(f"Row {idx + 2}: Validation error - {str(e)}")
                    except Exception as e:
                        errors.append(f"Row {idx + 2}: Unexpected error - {str(e)}")
                
                if errors:
                    console.print("[yellow]Validation warnings:[/yellow]")
                    for error in errors:
                        console.print(f"  [yellow]•[/yellow] {error}")
                
                progress.update(task, description=f"[green]Parsed {len(entries)} work log entry(ies)[/green]")
            
            return entries
            
        except Exception as e:
            console.print(f"[red]Error importing from Excel:[/red] {str(e)}")
            return []
    
    def update_excel_status(self, input_file: str, results: List[SyncResult]) -> bool:
        """Update Excel file with sync status.
        
        Args:
            input_file: Input Excel file path
            results: List of sync results
            
        Returns:
            True if successful, False otherwise
        """
        try:
            # Read existing Excel
            wb = load_workbook(input_file)
            ws = wb['Work Logs']
            
            # Create status mapping
            status_map = {r.issue_key: r for r in results}
            
            # Update status column (column G)
            for row in range(2, ws.max_row + 1):
                issue_key = ws[f'A{row}'].value
                if issue_key and str(issue_key).strip() in status_map:
                    result = status_map[str(issue_key).strip()]
                    status_cell = ws[f'G{row}']
                    if result.success:
                        status_cell.value = "✓ Synced"
                        status_cell.font = Font(color="00AA00")  # Green
                    else:
                        status_cell.value = f"✗ {result.message[:30]}"
                        status_cell.font = Font(color="FF0000")  # Red
            
            # Save updated file
            output_file = input_file.replace('.xlsx', '_synced.xlsx')
            wb.save(output_file)
            
            console.print(f"[green]✓[/green] Updated Excel file: [cyan]{output_file}[/cyan]")
            return True
            
        except Exception as e:
            console.print(f"[yellow]Warning: Could not update Excel status:[/yellow] {str(e)}")
            return False
    
    def export_worklog_summary(
        self, 
        worklogs: List[ExistingWorkLog], 
        issues_dict: dict, 
        output_file: str,
        hierarchical_groups: Optional[List] = None,
        all_issues: Optional[List] = None
    ) -> bool:
        """Export worklog summary to Excel with original values tracking.
        
        Args:
            worklogs: List of ExistingWorkLog objects
            issues_dict: Dictionary mapping issue keys to (summary, type) tuples
            output_file: Output Excel file path
            hierarchical_groups: Optional list of (epic_key, HierarchicalGroup) tuples for grouped export
            all_issues: Optional list of all Issue objects for hierarchy display
            
        Returns:
            True if successful, False otherwise
        """
        try:
            if not worklogs and not hierarchical_groups:
                console.print("[yellow]No work logs to export.[/yellow]")
                return False
            
            with Progress(
                SpinnerColumn(),
                TextColumn("[progress.description]{task.description}"),
                console=console
            ) as progress:
                task = progress.add_task("Creating worklog summary Excel...", total=None)
                
                # Create DataFrame from worklogs
                data = []
                
                if hierarchical_groups:
                    # Export with hierarchical grouping: Epic > Story/Task > Subtask (recursive tree view)
                    from ..services.hierarchy_service import HierarchicalGroup
                    
                    # Create issue map from all_issues to get parent information
                    issue_map = {}
                    if all_issues:
                        issue_map = {issue.key: issue for issue in all_issues}
                    
                    epic_counter = 0  # Track Epic numbers (1, 2, 3, ...)
                    
                    for epic_key, group in hierarchical_groups:
                        # Create subtasks map for this group for recursive traversal
                        subtasks_map = group.subtasks_map
                        epic_counter += 1
                        epic_number = str(epic_counter)  # Epic number: 1, 2, 3, ...
                        
                        # Build complete children map: map each parent issue key to all its children
                        # This includes both subtasks and tasks that are children of stories/tasks
                        children_map = defaultdict(list)
                        # Add subtasks from subtasks_map
                        for parent_key, children_list in subtasks_map.items():
                            children_map[parent_key].extend(children_list)
                        # Also check all_issues for tasks/stories that have children
                        if all_issues:
                            for issue in all_issues:
                                # If issue has a parent_key, it's a child of that parent
                                if issue.parent_key:
                                    if issue.parent_key not in children_map:
                                        children_map[issue.parent_key] = []
                                    if issue not in children_map[issue.parent_key]:
                                        children_map[issue.parent_key].append(issue)
                        
                        def dump_issue_tree_recursive(
                            issue: Issue, 
                            depth: int, 
                            number_prefix: str, 
                            parent_issue: Optional[Issue] = None, 
                            child_index: int = 0,
                            visited: Optional[set] = None,
                            max_depth: int = 20
                        ) -> None:
                            """Recursively dump issue tree to Excel (following recursion best practices).
                            
                            Filesystem analogy: recursively traverse directory tree with cycle detection.
                            
                            Implementation follows Recursion-of-Thought principles:
                            1. Base case: Issue has no children → process and return
                            2. Recursive case: Process issue, then recurse on each child
                            3. Cycle detection: Track visited issues to prevent infinite loops
                            4. Depth limiting: Prevent stack overflow for deeply nested structures
                            
                            Args:
                                issue: Issue node to process
                                depth: Current depth level (0=Epic/root, 1=Story/Task, 2+=Subtask)
                                number_prefix: Hierarchical number prefix (e.g., "1", "1.1", "1.1.1")
                                parent_issue: Parent issue node (for parent info extraction)
                                child_index: Index of this child among siblings (0-based)
                                visited: Set of visited issue keys (for cycle detection)
                                max_depth: Maximum recursion depth to prevent stack overflow
                            
                            Edge cases handled:
                            - Empty children list: Base case, just processes the issue
                            - Null/None issue: Should not occur (caller validates)
                            - Cycle detection: Skips if issue already visited
                            - Max depth exceeded: Stops recursion and logs warning
                            """
                            nonlocal data, issue_map, issues_dict, worklogs, subtasks_map, epic_number, children_map
                            
                            # Base case 1: Cycle detection - prevent infinite loops
                            if visited is None:
                                visited = set()
                            if issue.key in visited:
                                # Cycle detected - skip this node to prevent infinite recursion
                                return
                            visited.add(issue.key)
                            
                            # Base case 2: Depth limit - prevent stack overflow
                            if depth > max_depth:
                                # Depth limit exceeded - stop recursion to prevent stack overflow
                                # Note: We could log this, but it's an edge case that shouldn't occur in normal JIRA hierarchies
                                return
                            
                            # Step 1: Build hierarchy number for this node
                            # - Epics use epic_number (1, 2, 3, ...)
                            # - Children append their index+1 to parent's number (1.1, 1.2, 1.1.1, ...)
                            if depth == 0:
                                hierarchy_number = epic_number  # Root/Epic level
                            else:
                                hierarchy_number = f"{number_prefix}.{child_index + 1}" if number_prefix else str(child_index + 1)
                            
                            # Step 2: Extract issue information
                            issue_summary, issue_type = issues_dict.get(issue.key, (issue.summary or issue.key, issue.issue_type))
                            if not issue_summary or issue_summary.strip() == "":
                                issue_summary = issue.key  # Fallback to key if summary empty
                            
                            # Step 3: Build indentation prefix (2 spaces per depth level, root has no indentation)
                            # Indentation formula: "  " * max(0, depth - 1) + "└─ " for children
                            if depth == 0:
                                indentation = ""  # Root/Epic - no indentation
                            else:
                                indentation = "  " * (depth - 1) + "└─ "  # Children - 2 spaces per level
                            
                            # Step 4: Determine parent information
                            parent_key_str = ""
                            parent_type_str = ""
                            parent_indicator = ""
                            
                            if issue.parent_epic_key and not issue.parent_key:
                                # Direct child of Epic (Story or Task directly under Epic)
                                # This includes Tasks that are direct children of Epics
                                parent_key_str = issue.parent_epic_key
                                parent_type_str = issue.parent_issue_type or "Epic"
                                if depth == 1:  # Only show indicator for direct Epic children
                                    parent_indicator = f" ({parent_type_str})"
                            elif issue.parent_key and parent_issue:
                                # Child of Story/Task (e.g., Task under Story, Subtask under Task)
                                parent_key_str = parent_issue.key
                                parent_type_str = issue.parent_issue_type or parent_issue.issue_type
                            elif issue.parent_epic_key:
                                # Has Epic link but also has parent - parent is more specific
                                if parent_issue:
                                    parent_key_str = parent_issue.key
                                    parent_type_str = issue.parent_issue_type or parent_issue.issue_type
                            
                            # Step 5: Find worklogs for this issue
                            issue_worklogs = [wl for wl in worklogs if wl.issue_key == issue.key]
                            
                            # Step 6: Process current node (base case - single node processing)
                            if issue_worklogs:
                                # Node has worklogs - add each worklog as a row
                                for wl in issue_worklogs:
                                    summary_with_indicator = f"{indentation}{issue_summary}{parent_indicator}" if parent_indicator else f"{indentation}{issue_summary}"
                                    row = wl.to_excel_row(summary_with_indicator, issue_type, parent_key_str, parent_type_str, hierarchy_number)
                                    data.append(row)
                            else:
                                # Node has no worklogs - add single row with zero time
                                summary_with_indicator = f"{indentation}{issue_summary}{parent_indicator}" if parent_indicator else f"{indentation}{issue_summary}"
                                issue_row = {
                                    "Hierarchy Number": hierarchy_number,
                                    "Worklog ID": "",
                                    "Issue Key": issue.key,
                                    "Summary": summary_with_indicator,
                                    "Type": issue_type,
                                    "Parent Issue Key": parent_key_str,
                                    "Parent Issue Type": parent_type_str,
                                    "Time Logged (hours)": "0",
                                    "Original Time (hours)": "0",
                                    "Date": "",
                                    "Comment": "",
                                    "Original Comment": "",
                                    "Author": "",
                                    "Status": "No Worklog"
                                }
                                data.append(issue_row)
                            
                            # Step 7: Recursive case - process children (if any)
                            # Base case: If no children, recursion stops here (implicit return)
                            # Get children from complete children_map (includes all children: tasks, subtasks, etc.)
                            children = children_map.get(issue.key, [])
                            
                            if not children:
                                # Base case: No children - recursion terminates naturally
                                return
                            
                            # Recursive step: Process each child with increased depth and updated number prefix
                            # Filesystem analogy: Traverse subdirectories recursively
                            for child_index_num, child in enumerate(children):
                                # Recursive call with:
                                # - Increased depth (depth + 1)
                                # - Updated number prefix (current hierarchy_number)
                                # - Current issue as parent
                                # - Child index for numbering
                                # - Visited set (shared across recursion) for cycle detection
                                dump_issue_tree_recursive(
                                    issue=child,
                                    depth=depth + 1,
                                    number_prefix=hierarchy_number,
                                    parent_issue=issue,
                                    child_index=child_index_num,
                                    visited=visited,  # Share visited set to detect cross-tree cycles
                                    max_depth=max_depth
                                )
                        
                        # Wrapper: Process Epic root, then recursively dump its subtree
                        # Base case: If Epic doesn't exist, skip this group
                        if not group.epic:
                            continue
                        
                        # Step 1: Process Epic root node (depth 0)
                        epic_summary, epic_type = issues_dict.get(group.epic.key, (group.epic.summary or group.epic.key, group.epic.issue_type))
                        if not epic_summary or epic_summary.strip() == "":
                            epic_summary = group.epic.key
                        epic_row = {
                            "Hierarchy Number": epic_number,
                            "Worklog ID": "",
                            "Issue Key": group.epic.key,
                            "Summary": epic_summary,  # No indentation for root
                            "Type": epic_type,
                            "Parent Issue Key": "",
                            "Parent Issue Type": "",
                            "Time Logged (hours)": "",
                            "Original Time (hours)": "",
                            "Date": "",
                            "Comment": "",
                            "Original Comment": "",
                            "Author": "",
                            "Status": "Epic"
                        }
                        data.append(epic_row)
                        
                        # Step 2: Recursively dump all children (Stories/Tasks) under Epic
                        # Filesystem analogy: Traverse root directory contents
                        for story_index, story_task in enumerate(group.stories_tasks):
                            dump_issue_tree_recursive(
                                issue=story_task,
                                depth=1,  # Start at depth 1 (children of root)
                                number_prefix=epic_number,
                                parent_issue=group.epic,
                                child_index=story_index,
                                visited=set(),  # New visited set for each Epic subtree
                                max_depth=20
                            )
                        
                        # No empty row between groups - removed to avoid empty lines
                else:
                    # Export flat list (original behavior)
                    # Create issue map from all_issues to get parent information
                    issue_map = {}
                    if all_issues:
                        issue_map = {issue.key: issue for issue in all_issues}
                    
                    for wl in worklogs:
                        issue_key = wl.issue_key
                        summary, issue_type = issues_dict.get(issue_key, ("", ""))
                        # Use issue key as fallback if summary is empty
                        if not summary or summary.strip() == "":
                            summary = issue_key
                        
                        # Get parent information
                        parent_key_str = ""
                        parent_type_str = ""
                        if issue_key in issue_map:
                            issue = issue_map[issue_key]
                            if issue.parent_epic_key:
                                parent_key_str = issue.parent_epic_key
                                parent_type_str = issue.parent_issue_type or "Epic"
                            elif issue.parent_key:
                                parent_key_str = issue.parent_key
                                parent_type_str = issue.parent_issue_type or ""
                                if not parent_type_str and parent_key_str in issue_map:
                                    parent_type_str = issue_map[parent_key_str].issue_type
                        
                        # Flat list doesn't have hierarchy numbers
                        data.append(wl.to_excel_row(summary, issue_type, parent_key_str, parent_type_str, ""))
                
                df = pd.DataFrame(data)
                
                # Define column order for worklog summary
                worklog_columns = [
                    "Hierarchy Number",
                    "Worklog ID",
                    "Issue Key",
                    "Summary",
                    "Type",
                    "Parent Issue Key",
                    "Parent Issue Type",
                    "Time Logged (hours)",
                    "Original Time (hours)",
                    "Date",
                    "Comment",
                    "Original Comment",
                    "Author",
                    "Status"
                ]
                
                # Ensure all columns exist
                for col in worklog_columns:
                    if col not in df.columns:
                        df[col] = ""
                
                # Reorder columns
                df = df[worklog_columns]
                
                progress.update(task, description="Writing Excel file...")
                
                # Write to Excel with formatting
                output_path = Path(output_file)
                output_path.parent.mkdir(parents=True, exist_ok=True)
                
                # Create Excel writer
                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Worklog Summary')
                    
                    # Get workbook and worksheet for formatting
                    workbook = writer.book
                    worksheet = writer.sheets['Worklog Summary']
                    
                    # Format header row
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF")
                    
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Format columns
                    worksheet.column_dimensions['A'].width = 15  # Worklog ID
                    worksheet.column_dimensions['B'].width = 15  # Issue Key
                    worksheet.column_dimensions['C'].width = 50  # Summary
                    worksheet.column_dimensions['D'].width = 15  # Type
                    worksheet.column_dimensions['E'].width = 18  # Time Logged
                    worksheet.column_dimensions['F'].width = 20  # Original Time
                    worksheet.column_dimensions['G'].width = 15  # Date
                    worksheet.column_dimensions['H'].width = 40  # Comment
                    worksheet.column_dimensions['I'].width = 40  # Original Comment
                    worksheet.column_dimensions['J'].width = 20  # Author
                    worksheet.column_dimensions['K'].width = 15  # Status
                    
                    # Format data cells
                    for row in range(2, len(df) + 2):
                        # Worklog ID - read-only indicator
                        wl_id_cell = worksheet[f'A{row}']
                        if wl_id_cell.value:
                            wl_id_cell.font = Font(color="808080")  # Gray for read-only
                        
                        # Issue Key - read-only indicator
                        issue_key_cell = worksheet[f'B{row}']
                        if issue_key_cell.value:
                            issue_key_cell.font = Font(color="808080")  # Gray for read-only
                        
                        # Original Time - read-only indicator (gray background)
                        orig_time_cell = worksheet[f'F{row}']
                        if orig_time_cell.value:
                            orig_time_cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                            orig_time_cell.font = Font(color="808080")
                        
                        # Original Comment - read-only indicator (gray background)
                        orig_comment_cell = worksheet[f'I{row}']
                        if orig_comment_cell.value:
                            orig_comment_cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                            orig_comment_cell.font = Font(color="808080")
                        
                        # Time Logged - editable (decimal format)
                        time_cell = worksheet[f'E{row}']
                        time_cell.number_format = '0.00'
                        
                        # Date - date format
                        date_cell = worksheet[f'G{row}']
                        date_cell.number_format = 'YYYY-MM-DD'
                    
                    # Add summary row at the end
                    summary_row = len(df) + 2
                    worksheet[f'A{summary_row}'] = "TOTAL"
                    worksheet[f'A{summary_row}'].font = Font(bold=True, color="FFFFFF")
                    worksheet[f'A{summary_row}'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    worksheet[f'A{summary_row}'].alignment = Alignment(horizontal="right", vertical="center")
                    
                    # Sum formula for time logged
                    worksheet[f'E{summary_row}'] = f"=SUM(E2:E{len(df) + 1})"
                    worksheet[f'E{summary_row}'].number_format = '0.00'
                    worksheet[f'E{summary_row}'].font = Font(bold=True, color="FFFFFF")
                    worksheet[f'E{summary_row}'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    
                    # Sum formula for original time
                    worksheet[f'F{summary_row}'] = f"=SUM(F2:F{len(df) + 1})"
                    worksheet[f'F{summary_row}'].number_format = '0.00'
                    worksheet[f'F{summary_row}'].font = Font(bold=True, color="FFFFFF")
                    worksheet[f'F{summary_row}'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    
                    # Merge cells for total label if needed (optional)
                    # worksheet.merge_cells(f'A{summary_row}:D{summary_row}')
                    
                    # Add instructions sheet
                    instructions_df = pd.DataFrame({
                        'Column': [
                            'Worklog ID',
                            'Issue Key',
                            'Summary',
                            'Type',
                            'Time Logged (hours)',
                            'Original Time (hours)',
                            'Date',
                            'Comment',
                            'Original Comment',
                            'Author',
                            'Status'
                        ],
                        'Description': [
                            'Jira worklog ID (read-only)',
                            'Jira issue key (read-only)',
                            'Issue summary (read-only)',
                            'Issue type (read-only)',
                            'Time logged in hours - EDIT THIS (decimal, e.g., 2.5)',
                            'Original time logged (read-only, gray background)',
                            'Work log date (YYYY-MM-DD)',
                            'Work log comment - EDIT THIS',
                            'Original comment (read-only, gray background)',
                            'Work log author (read-only)',
                            'Sync status (auto-populated)'
                        ],
                        'Editable': [
                            'No',
                            'No',
                            'No',
                            'No',
                            'Yes',
                            'No',
                            'Yes',
                            'Yes',
                            'No',
                            'No',
                            'No'
                        ]
                    })
                    
                    instructions_df.to_excel(writer, index=False, sheet_name='Instructions')
                    
                    # Format instructions sheet
                    inst_worksheet = writer.sheets['Instructions']
                    for cell in inst_worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    inst_worksheet.column_dimensions['A'].width = 25
                    inst_worksheet.column_dimensions['B'].width = 60
                    inst_worksheet.column_dimensions['C'].width = 15
                
                progress.update(task, description=f"[green]Worklog summary Excel created: {output_file}[/green]")
            
            console.print(f"[green]✓[/green] Exported {len(worklogs)} work log(s) to [cyan]{output_file}[/cyan]")
            console.print(f"[dim]You can edit 'Time Logged (hours)' and 'Comment' columns to update work logs.[/dim]")
            console.print(f"[dim]Original values are preserved in gray columns for reference.[/dim]")
            return True
            
        except Exception as e:
            console.print(f"[red]Error exporting worklog summary to Excel:[/red] {str(e)}")
            return False
    
    def import_worklog_summary_diff(self, input_file: str) -> List[WorkLogUpdate]:
        """Import worklog summary from Excel and detect changes (diff).
        
        Args:
            input_file: Input Excel file path
            
        Returns:
            List of WorkLogUpdate objects with detected changes
        """
        try:
            input_path = Path(input_file)
            if not input_path.exists():
                raise FileNotFoundError(f"Excel file not found: {input_file}")
            
            with Progress(
                SpinnerColumn(),
                TextColumn("[progress.description]{task.description}"),
                console=console
            ) as progress:
                task = progress.add_task("Reading worklog summary Excel...", total=None)
                
                # Read Excel file
                df = pd.read_excel(input_path, sheet_name='Worklog Summary')
                
                progress.update(task, description="Detecting changes...")
                
                # Validate required columns
                required_columns = ['Worklog ID', 'Issue Key', 'Time Logged (hours)', 'Original Time (hours)', 'Date']
                missing_columns = [col for col in required_columns if col not in df.columns]
                if missing_columns:
                    raise ValueError(f"Missing required columns: {', '.join(missing_columns)}")
                
                # Parse and detect changes
                updates = []
                errors = []
                
                for idx, row in df.iterrows():
                    try:
                        # Get values
                        worklog_id = str(row['Worklog ID']).strip() if pd.notna(row['Worklog ID']) else ""
                        issue_key = str(row['Issue Key']).strip() if pd.notna(row['Issue Key']) else ""
                        time_str = str(row['Time Logged (hours)']).strip() if pd.notna(row['Time Logged (hours)']) else ""
                        orig_time_str = str(row['Original Time (hours)']).strip() if pd.notna(row['Original Time (hours)']) else ""
                        date_str = str(row['Date']).strip() if pd.notna(row['Date']) else ""
                        comment = str(row.get('Comment', '')).strip() if pd.notna(row.get('Comment', '')) else ""
                        orig_comment = str(row.get('Original Comment', '')).strip() if pd.notna(row.get('Original Comment', '')) else ""
                        
                        # Skip rows with empty required fields
                        if not worklog_id or not issue_key or not time_str or not orig_time_str:
                            continue
                        
                        # Validate issue key
                        if not validate_issue_key(issue_key):
                            errors.append(f"Row {idx + 2}: Invalid issue key format: {issue_key}")
                            continue
                        
                        # Parse times
                        try:
                            new_time_hours = parse_time_hours(time_str)
                            original_time_hours = parse_time_hours(orig_time_str)
                        except ValueError as e:
                            errors.append(f"Row {idx + 2}: {str(e)}")
                            continue
                        
                        # Parse date
                        try:
                            if isinstance(row['Date'], pd.Timestamp):
                                work_date = row['Date'].date()
                            elif isinstance(row['Date'], str):
                                work_date = parse_date(date_str)
                            else:
                                work_date = parse_date(date_str)
                        except ValueError as e:
                            errors.append(f"Row {idx + 2}: {str(e)}")
                            continue
                        
                        # Create work log update (even if no changes detected)
                        # Using alias 'date' for Excel column compatibility
                        update = WorkLogUpdate(
                            worklog_id=worklog_id,
                            issue_key=issue_key,
                            original_time_hours=original_time_hours,
                            new_time_hours=new_time_hours,
                            original_comment=orig_comment if orig_comment else None,
                            new_comment=comment if comment else None,
                            date=work_date  # Using alias
                        )
                        
                        # Only add if there are changes
                        if update.has_changes():
                            updates.append(update)
                        
                    except ValidationError as e:
                        errors.append(f"Row {idx + 2}: Validation error - {str(e)}")
                    except Exception as e:
                        errors.append(f"Row {idx + 2}: Unexpected error - {str(e)}")
                
                if errors:
                    console.print("[yellow]Validation warnings:[/yellow]")
                    for error in errors:
                        console.print(f"  [yellow]•[/yellow] {error}")
                
                if updates:
                    progress.update(task, description=f"[green]Detected {len(updates)} change(s)[/green]")
                else:
                    progress.update(task, description="[yellow]No changes detected[/yellow]")
            
            return updates
            
        except Exception as e:
            console.print(f"[red]Error importing worklog summary diff:[/red] {str(e)}")
            return []

